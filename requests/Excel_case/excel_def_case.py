#coding=utf-8
import openpyxl
import os
base_bash = os.getcwd()
latest_bash = os.path.dirname(base_bash)

class HandExcel:
    def open_excel(self):
        '''
        :return: excel文件
        '''
        excel = openpyxl.load_workbook(latest_bash+"/Case/imooc_test.xlsx")
        return excel


    def get_sheetnames(self):
        '''
        :return: 所有sheet页名称，同时也是获取所有sheet页中的对象
        '''
        sheet_name = self.open_excel().sheetnames
        return sheet_name


    def get_sheetname_data(self,number=None):
        '''
        获取某一个sheet页名称
        :param number: 第几个sheet页，从0开始，如：0
        :return: 返回某一个sheet页的名称，如返回第一个sheet名称：sheet1，同时也是获取单个sheet中的对象
                可以通过这个对象获取对应的value数据
        '''
        if number == None:
            number = 0
        sheet_name = self.open_excel().sheetnames
        data = self.open_excel()[sheet_name[number]]
        return data

    def get_sheet_value(self,row,column):
        '''
        获取某一个单元格的内容
        :param row:行数，从1开始，例如：5
        :param column:列数，从1开始，例如：6
        :return:返回第row行第column列单元格的内容
        '''
        sheet_value = self.get_sheetname_data().cell(row,column).value
        return sheet_value

    def get_row_max(self):
        '''
        :return: 单个sheet表中最大行数
        '''
        row_max = self.get_sheetname_data().max_row
        return row_max

    def get_column_max(self):
        '''
        :return: 单个sheet表中最大列数
        '''
        column_max = self.get_sheetname_data().max_column
        return column_max

    def get_row_value(self,row):
        '''
        1、获取某一行的内容
        2、通过get_sheetname_data()方法，获取某一行的对象，通过append()将单个单元格内容写入row_list中
        :param row: 行号，如：1
        :return: 返回list数据，返回第row行的内容
        '''
        row_list = []
        row_data = self.get_sheetname_data()[row]
        for row_object in row_data:
            row_list.append(row_object.value)
        return row_list

Handexcel = HandExcel()

if __name__ == '__main__':
    handexcel = HandExcel()
    data = handexcel.get_row_value(4)
    print(data)