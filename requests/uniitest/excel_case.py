#coding=utf-8
'''
1、通过openpyxl模块，打开excel文件
2、通过openpyxl.load_workbook()函数获取文件路径
3、excel_path.sheetnames获取所有的sheet表格名称
4、excel_path.sheetnames[0]获取第一个sheet名称
'''
import os
import sys
base_path = os.getcwd()
latest_path = os.path.dirname(base_path)
sys.path.append(base_path)
import openpyxl
excel_path = openpyxl.load_workbook(latest_path+"/Case/imooc_test.xlsx")
sheet_names = excel_path.sheetnames
sheet_name = excel_path.sheetnames[0]
sheet_name.cell()
print(sheet_names)
print(sheet_name)
